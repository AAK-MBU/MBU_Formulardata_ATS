""" Script to upload fetch an OS2-formular submission and upload it in pdf format to Sharepoint. """
from urllib.parse import unquote, urlparse

import json
import urllib.parse

from io import BytesIO

import math

from typing import Dict, Any

import requests

import pandas as pd

from sqlalchemy import create_engine

from openpyxl.styles import Alignment

from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection

from mbu_dev_shared_components.msoffice365.sharepoint_api.files import Sharepoint


### REMOVE THESE IMPORTS
from typing import Optional, List

from openpyxl.styles import Font
from openpyxl import load_workbook
### REMOVE THESE IMPORTS


### REMOVE THESE 2 FUNCTIONS AFTER UPDATING mbu-dev-shared-components
def append_row_to_sharepoint_excel(
    sharepoint: Sharepoint,
    folder_name: str = "",
    excel_file_name: str = "",
    sheet_name: str = "",
    new_row: Dict = None,
) -> None:
    """
    • Appends a row to an existing Excel file.
    • Sorts and formats based on provided parameters.
    """

    # 1. Pull file
    binary_file = sharepoint.fetch_file_using_open_binary(excel_file_name, folder_name)

    if binary_file is None:
        raise FileNotFoundError(f"File '{excel_file_name}' not found in folder '{folder_name}'.")

    wb = load_workbook(BytesIO(binary_file))

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{excel_file_name}'")

    ws = wb[sheet_name]

    # 2 Clean up empty rows before appending
    for row_idx in range(ws.max_row, 1, -1):  # Start from bottom, skip header
        row_values = [cell.value for cell in ws[row_idx]]

        if all(cell is None for cell in row_values):
            ws.delete_rows(row_idx)

    # 3. Append new row to sheet
    ws.append([new_row.get(header.value, "") for header in ws[1]])

    # 4. Save and upload
    temp_stream = BytesIO()

    wb.save(temp_stream)

    temp_stream.seek(0)

    sharepoint.upload_file_from_bytes(temp_stream.getvalue(), excel_file_name, folder_name)

    print(f"✔ Added row + sorted '{sheet_name}' in '{excel_file_name}'.")


def format_and_sort_excel_file(
    sharepoint: Sharepoint,
    folder_name: str,
    excel_file_name: str,
    sheet_name: str,
    sorting_keys: Optional[List[Dict[str, Any]]] = None,
    font_config: Optional[Dict[int, Dict[str, Any]]] = None,
    bold_rows: Optional[List[int]] = None,
    italic_rows: Optional[List[int]] = None,
    align_horizontal: str = "center",
    align_vertical: str = "center",
    column_widths: Any = "auto",
    freeze_panes: Optional[str] = None,
):
    """
    Sorts and formats an Excel worksheet based on provided styling and sorting rules.

    Params:
        folder_name: Name of the folder where the file resides
        excel_file_name: Name of the excel file
        sheet_name: Name of the sheet that will be sorted
        sorting_keys: List of dicts like [{"key": "A", "ascending": True, "type": "datetime"}]
        font_config: Dict of row -> font config (overrides bold/italic)
        bold_rows: List of row numbers to bold like [1, 2, ...]
        italic_rows: List of row numbers to italicize like [1, 2, ...]
        align_horizontal: Horizontal text alignment
        align_vertical: Vertical text alignment
        column_widths: "auto" or an int to represent a pixel value
        freeze_panes: E.g., "A2" to freeze header row

    Returns:
        Modified worksheet
    """

    # Step 1 - Fetch the file to update from SharePoint and load it as a workbook
    # This ensures we don't override any other sheets in the excel file
    binary_file = sharepoint.fetch_file_using_open_binary(excel_file_name, folder_name)
    if binary_file is None:
        raise FileNotFoundError(f"File '{excel_file_name}' not found in folder '{folder_name}'.")

    wb = load_workbook(BytesIO(binary_file))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{excel_file_name}'")

    ws = wb[sheet_name]

    # Step 2 - Read data into DataFrame
    rows = list(ws.iter_rows(values_only=True))
    header, *data_rows = rows
    df = pd.DataFrame(data_rows, columns=header)

    # Step 3 – Prepare sorting logic
    # For each sorting instruction, we:
    # - Extract the column to sort by (using letter, index, or name)
    # - Convert the column values to the desired data type if specified (str, int, float, datetime)
    # - Track which columns to sort and in which order (ascending or descending)
    #
    # This ensures the DataFrame is sorted correctly, even when types like dates or numbers need conversion.
    if sorting_keys:
        sort_columns = []
        ascending_flags = []

        for item in sorting_keys:
            key = item.get("key")
            ascending = item.get("ascending", True)
            dtype = item.get("type")

            if isinstance(key, int):
                col_name = header[key]

            elif isinstance(key, str) and key.isalpha():
                col_name = header[ord(key.upper()) - ord("A")]

            else:
                col_name = key

            sort_columns.append(col_name)
            ascending_flags.append(ascending)

            if dtype == "datetime":
                df[col_name] = pd.to_datetime(df[col_name], dayfirst=True, errors="coerce")

            elif dtype == "int":
                df[col_name] = pd.to_numeric(df[col_name], errors="coerce", downcast="integer")

            elif dtype == "float":
                df[col_name] = pd.to_numeric(df[col_name], errors="coerce", downcast="float")

            elif dtype == "str":
                df[col_name] = df[col_name].astype(str)

        # Step 4 – Sort
        df.sort_values(by=sort_columns, ascending=ascending_flags, inplace=True)

    # Step 5 - Overwrite worksheet
    ws.delete_rows(1, ws.max_row)

    ws.append(header)

    for _, row in df.iterrows():
        ws.append(list(row))

    # Step 6 – Adjust column widths and apply wrapping if needed
    #
    # If column_widths is "auto":
    # - Calculate the max content length in each column and set the column width accordingly (+2 for padding)
    #
    # If column_widths is a single int:
    # - Use it as a global max width across all columns
    # - If content fits, set width based on actual content length
    # - If content exceeds the max width clamp column width and enable wrap_text for that column's cells
    #
    # Then, for wrapped cells, auto-adjust the row height:
    # - Estimate how many lines the wrapped text would occupy and set row height accordingly to ensure all content is visible
    if column_widths in (None, "auto"):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)

            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    elif isinstance(column_widths, int):
        for col in ws.columns:
            col_letter = col[0].column_letter

            max_len = max(len(str(cell.value or "")) for cell in col)

            # If content fits, auto-size
            if max_len + 2 <= column_widths:
                ws.column_dimensions[col_letter].width = max_len + 2

            # Else, cap width and enable wrap
            else:
                ws.column_dimensions[col_letter].width = column_widths

                for cell in col:
                    cell.alignment = Alignment(wrap_text=True)

        # Here we handle row height
        for row in ws.iter_rows():
            max_line_count = 1

            for cell in row:
                if cell.value and cell.alignment and cell.alignment.wrap_text:
                    col_letter = cell.column_letter
                    col_width = ws.column_dimensions[col_letter].width or 10
                    chars_per_line = col_width * 1.2
                    lines = str(cell.value).split("\n")
                    line_count = sum(math.ceil(len(line) / chars_per_line) for line in lines)
                    max_line_count = max(max_line_count, line_count)

            ws.row_dimensions[row[0].row].height = max_line_count * 20

    else:
        raise ValueError(f"Column width provided with incorrect datatype - datatype int expected, instead column width is of datatype {type(column_widths)}")

    # Step 7 - Freeze panes if needed
    if freeze_panes:
        ws.freeze_panes = freeze_panes

    # Step 8 – Apply base formatting
    # For each cell in the worksheet:
    # - Apply font styling based on either a custom `font_config` (row-specific) or default to bold/italic based on row number (e.g., header rows)
    # - Set horizontal and vertical alignment for consistent layout
    # - Disable text wrapping by default (wrapping will be handled later if needed)
    #
    # This ensures a clean, uniform look across the sheet while allowing for custom styling where defined.
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if font_config and row_idx in font_config:
                config = font_config[row_idx]

                cell.font = Font(
                    name=config.get("name", "Calibri"),
                    size=config.get("size", 11),
                    bold=config.get("bold", False),
                    italic=config.get("italic", False),
                )

            else:
                cell.font = Font(
                    bold=row_idx in bold_rows if bold_rows else False,
                    italic=row_idx in italic_rows if italic_rows else False,
                )

            cell.alignment = Alignment(
                horizontal=align_horizontal,
                vertical=align_vertical,
                wrap_text=cell.alignment.wrap_text
            )

    # Step 9 - Save and re-upload
    temp_stream = BytesIO()

    wb.save(temp_stream)

    temp_stream.seek(0)

    sharepoint.upload_file_from_bytes(temp_stream.getvalue(), excel_file_name, folder_name)
### REMOVE THESE 2 FUNCTIONS AFTER UPDATING mbu-dev-shared-components


def load_credential(url, token, credential_name: str) -> dict:
    """
    Fetch a credential object from the Automation Server using its name,
    and return the full credential dictionary with 'data' parsed into a dict.
    """

    if not url or not token:
        raise EnvironmentError("ATS_URL or ATS_TOKEN is not set in the environment")

    headers = {
        "Authorization": f"Bearer {token}"
    }

    full_url = f"{url}/credentials/by_name/{credential_name}"

    response = requests.get(full_url, headers=headers, timeout=60)

    response.raise_for_status()

    credential = response.json()

    # Parse the JSON string stored in the 'data' field
    try:
        credential["data"] = json.loads(credential.get("data", "{}"))

    except json.JSONDecodeError:
        credential["data"] = {}
        print(f"Warning: Failed to decode 'data' for credential '{credential_name}'.")

    return credential


def get_credentials_and_constants(orchestrator_connection: OrchestratorConnection) -> Dict[str, Any]:
    """Retrieve necessary credentials and constants from the orchestrator connection."""
    try:
        credentials = {
            "go_api_endpoint": orchestrator_connection.get_constant('go_api_endpoint').value,
            "go_api_username": orchestrator_connection.get_credential('go_api').username,
            "go_api_password": orchestrator_connection.get_credential('go_api').password,
            "os2_api_key": orchestrator_connection.get_credential('os2_api').password,
            "sql_conn_string": orchestrator_connection.get_constant('DbConnectionString').value,
            "journalizing_tmp_path": orchestrator_connection.get_constant('journalizing_tmp_path').value,
        }
        return credentials
    except AttributeError as e:
        raise SystemExit(e) from e


def get_workqueue_items(url, token, workqueue_id):
    """
    Retrieve items from the specified workqueue.
    If the queue is empty, return an empty list.
    """

    workqueue_items = set()

    if not url or not token:
        raise EnvironmentError("ATS_URL or ATS_TOKEN is not set in the environment")

    headers = {
        "Authorization": f"Bearer {token}"
    }

    full_url = f"{url}/workqueues/{workqueue_id}/items"

    response = requests.get(full_url, headers=headers, timeout=60)

    res_json = response.json().get("items", [])

    for row in res_json:
        ref = row.get("reference")

        workqueue_items.add(ref)

    return workqueue_items


def get_forms_data(conn_string: str, form_type: str) -> list[dict]:
    """
    Retrieve form_data['data'] for all matching submissions for the given form type,
    excluding purged entries.
    """

    query = """
        SELECT
            form_id,
            form_data,
            CAST(form_submitted_date AS datetime) AS form_submitted_date
        FROM
            [RPA].[journalizing].[Forms]
        WHERE
            form_type = ?
            AND form_data IS NOT NULL
            AND form_submitted_date IS NOT NULL
        ORDER BY form_submitted_date DESC
    """

    # Create SQLAlchemy engine
    encoded_conn_str = urllib.parse.quote_plus(conn_string)
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={encoded_conn_str}")

    try:
        df = pd.read_sql(sql=query, con=engine, params=(form_type,))

    except Exception as e:
        print("Error during pd.read_sql:", e)

        raise

    if df.empty:
        print("No submissions found for the given form type.")

        return []

    extracted_data = []
    for _, row in df.iterrows():
        try:
            parsed = json.loads(row["form_data"])
            if "purged" not in parsed:  # Skip purged entries
                extracted_data.append(parsed)

        except json.JSONDecodeError:
            print("Invalid JSON in form_data, skipping row.")

    return extracted_data


def upload_pdf_to_sharepoint(
    sharepoint_api: Sharepoint,
    folder_name: str,
    os2_api_key: str,
    file_url: str,
) -> None:
    """Main function to upload a PDF to Sharepoint."""

    print("Upload PDF to Sharepoint started.")

    existing_pdfs_sum = 0

    existing_pdfs = sharepoint_api.fetch_files_list(folder_name=folder_name)

    if existing_pdfs:
        existing_pdf_names = {file["Name"] for file in existing_pdfs}

    else:
        existing_pdf_names = set()

    path = urlparse(file_url).path
    filename = path.split('/')[-1]
    final_filename = f"{unquote(filename)}"

    print(file_url)
    print(final_filename)

    if final_filename in existing_pdf_names:
        print(f"File {final_filename} already exists in Sharepoint. Skipping download.")

        return

    print("Downloading PDF from OS2Forms API.")
    try:
        downloaded_file = download_file_bytes(file_url, os2_api_key)

    except requests.RequestException as error:
        print(f"Failed to download file: {error}")

    # Upload the file to Sharepoint
    sharepoint_api.upload_file_from_bytes(
        binary_content=downloaded_file,
        file_name=final_filename,
        folder_name=folder_name
    )


def download_file_bytes(url: str, os2_api_key: str) -> bytes:
    """Downloads the content of a file from a specified URL, appending an API key to the URL for authorization.
    The API key is retrieved from an environment variable 'OS2ApiKey'.

    Parameters:
    url (str): The URL from which the file will be downloaded.
    os2_api_key (str): The API-key for OS2Forms api.

    Returns:
    bytes: The content of the file as a byte stream.

    Raises:
    requests.RequestException: If the HTTP request fails for any reason.
    """

    headers = {
        'Content-Type': 'application/json',
        'api-key': f'{os2_api_key}'
    }

    response = requests.request(method='GET', url=url, headers=headers, timeout=60)

    response.raise_for_status()

    return response.content
